'''
This script cleans an Excel workbook by removing specific HTML tags 
(<span class="baixa"> and </span>) from the "Producte" column in all sheets, 
while preserving all other data and sheets.
'''

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# File paths
file_path = 'productes-beneficiari-2014-04-07.xlsx'
output_path = 'cleaned_productes-beneficiari.xlsx'

# Specify the sheet to process
sheet_to_check = 'Productes per beneficiari'  # Replace with the sheet name you want to process
column_name = 'Producte'  # Replace with the column name to clean

# Load the workbook using openpyxl
wb = load_workbook(file_path)
sheet = wb[sheet_to_check]

# Preserve merged cells information
merged_cells = list(sheet.merged_cells)

# Load the specific sheet to a DataFrame
df = pd.read_excel(file_path, sheet_name=sheet_to_check)

# Clean the specified column if it exists
if column_name in df.columns:
    df[column_name] = df[column_name].str.replace(r'<span class="baixa">', '', regex=True)
    df[column_name] = df[column_name].str.replace(r'</span>', '', regex=True)

# Save the cleaned sheet back to the workbook
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    # Write all sheets to the new workbook
    for sheet_name in wb.sheetnames:
        if sheet_name == sheet_to_check:
            # Replace the specified sheet with the cleaned DataFrame
            df.to_excel(writer, index=False, sheet_name=sheet_name)
        else:
            # Copy the original sheet for all others
            original_df = pd.read_excel(file_path, sheet_name=sheet_name)
            original_df.to_excel(writer, index=False, sheet_name=sheet_name)

# Reload the workbook to restore merged cells
cleaned_wb = load_workbook(output_path)
cleaned_sheet = cleaned_wb[sheet_to_check]

# Restore merged cells and align text to the top
for merged_cell in merged_cells:
    cleaned_sheet.merge_cells(str(merged_cell))
    # Align the text to the top for the first cell in the merged range
    top_left_cell = cleaned_sheet.cell(merged_cell.min_row, merged_cell.min_col)
    top_left_cell.alignment = Alignment(vertical='top')

# Restore merged cells
for merged_cell in merged_cells:
    cleaned_sheet.merge_cells(str(merged_cell))

# Save the workbook with merged cells restored
cleaned_wb.save(output_path)

print(f"File saved as {output_path} with {sheet_to_check} cleaned and merged cells preserved.")
